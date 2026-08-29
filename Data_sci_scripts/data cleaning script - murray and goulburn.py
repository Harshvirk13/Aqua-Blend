"""
Clean raw river/reservoir water-quality exports (.xlsx or .csv) into the
Supabase target schema:

    site_id, source_name, measurement_datetime, PH Value(PH),
    Turbidity(NTU), Water Temperature(C), Nitrogen as NOx(mg/L),
    TSS(mg/L), Colour(PCU)

measurement_datetime is written as YYYY/MM/DD (date only, no time),
matching the reference schema file
(401201_MURRAY_JINGELLIC_river_water_quality_final.csv).

Handles TWO different raw shapes automatically:

1. WIDE format (e.g. O_shannassy_river.xlsx): one row per site+date,
   one column per parameter already spread out.

2. LONG/TIDY format (e.g. WMIS_Batch2.csv): one row per
   site+parameter+date, with a 'parameter' column naming the
   measurement and a 'value' column holding the number. This is
   pivoted into the wide target shape automatically.

WHY THIS SCRIPT EXISTS (the bugs it fixes)
-------------------------------------------
1. Excel day/month swap (wide xlsx sources): dates were originally
   entered as text in Australian dd/mm/yyyy format. Any date whose
   DAY was <= 12 got silently reinterpreted by Excel as mm/dd/yyyy
   and stored as a real datetime object with month and day SWAPPED.
   Dates whose day was > 12 couldn't be misread as a month, so they
   were left alone as plain text in the correct dd/mm/yyyy order.
   This script detects which case each cell is in and reconstructs
   the true calendar date for both.

2. Mixed date encodings within one long CSV (e.g. WMIS_Batch2.csv):
   some stations store the date as a plain Excel serial number
   (e.g. "44927", days since 1899-12-30) while others store it as a
   US-style "M/D/YY" text string (confirmed by checking that dates
   increment day-by-day within a month, e.g. 1/1/23 ... 1/9/23,
   1/10/23, 1/11/23 - which only makes sense as month/day, not
   day/month). Both encodings are detected per-cell and converted.

USAGE
-----
    python3 clean_water_quality.py INPUT.xlsx [--sheet SHEET_NAME] [--out OUTPUT.csv]
    python3 clean_water_quality.py INPUT.csv [--out OUTPUT.csv]

If --sheet is omitted (xlsx only), the first sheet is used.
If --out is omitted, the output is written next to the input as
<input_stem>_cleaned.csv
"""

import argparse
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

EXCEL_EPOCH = datetime(1899, 12, 30)

TARGET_COLUMNS = [
    "site_id",
    "source_name",
    "measurement_datetime",
    "PH Value(PH)",
    "Turbidity(NTU)",
    "Water Temperature(C)",
    "Nitrogen as NOx(mg/L)",
    "TSS(mg/L)",
    "Colour(PCU)",
]

# Maps messy/variant source header names -> target schema column names.
# Add more aliases here as new source files show up with different spellings.
HEADER_ALIASES = {
    "turbity(ntu)": "Turbidity(NTU)",
    "turbidity(ntu)": "Turbidity(NTU)",
    "ph value(ph)": "PH Value(PH)",
    "water temperature(c)": "Water Temperature(C)",
    "nitrogen as nox(mg/l)": "Nitrogen as NOx(mg/L)",
    "tss(mg/l)": "TSS(mg/L)",
    "colour(pcu)": "Colour(PCU)",
    "site_id": "site_id",
    "source_name": "source_name",
    "measurement_datetime": "measurement_datetime",
    # long/tidy format source headers (e.g. WMIS exports)
    "station": "site_id",
    "station_name": "source_name",
    "datetime": "measurement_datetime",
    # "bare wide" format source headers (date + parameter columns, no
    # site_id/source_name in the file at all - e.g. *_wide.csv exports)
    "ph": "PH Value(PH)",
    "turbidity (ntu)": "Turbidity(NTU)",
    "water temperature (c)": "Water Temperature(C)",
    "total suspended solids (mg/l)": "TSS(mg/L)",
}

# Maps a source 'parameter' value (long/tidy format) -> target schema
# column name. Add more aliases as new parameter names show up.
# Parameters with no entry here (e.g. Streamflow, Salinity (EC),
# Stream Water Level, Alkalinity) are dropped, since the target
# schema has no matching column.
PARAMETER_ALIASES = {
    "ph": "PH Value(PH)",
    "turbidity": "Turbidity(NTU)",
    "water temperature": "Water Temperature(C)",
    "nitrogen as nox": "Nitrogen as NOx(mg/L)",
    "tss": "TSS(mg/L)",
    "colour": "Colour(PCU)",
}


def fix_datetime(val, dayfirst=True):
    """Return a corrected datetime.datetime (date-level only), or None.

    Handles, in order:
      - real datetime objects (from xlsx) with Excel's day/month swap undone
      - pure-digit strings/numbers -> Excel serial date (origin 1899-12-30)
      - 'D/M/Y' or 'M/D/Y' text strings, per `dayfirst`
    """
    if val is None:
        return None

    if isinstance(val, datetime):
        # Excel silently swapped day/month for any date with day <= 12
        # when it auto-parsed a dd/mm/yyyy text value as mm/dd/yyyy.
        # Swap back to recover the real date.
        return datetime(val.year, val.day, val.month)

    if isinstance(val, (int, float)):
        return EXCEL_EPOCH + timedelta(days=int(val))

    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        if s.isdigit():
            return EXCEL_EPOCH + timedelta(days=int(s))
        date_part = s.split(" ")[0]
        sep = "/" if "/" in date_part else "-"
        a, b, y = date_part.split(sep)
        y = int(y)
        if y < 100:
            y += 2000 if y < 70 else 1900
        d, m = (int(a), int(b)) if dayfirst else (int(b), int(a))
        return datetime(y, m, d)

    return None


def load_raw_rows(path: Path, sheet_name: str | None):
    if path.suffix.lower() == ".csv":
        # utf-8-sig quietly strips a BOM if present, common in WMIS exports
        raw = pd.read_csv(path, encoding="utf-8-sig", low_memory=False, dtype=str)
        header = list(raw.columns)
        data_rows = list(raw.itertuples(index=False, name=None))
        return header, data_rows

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]
    return header, data_rows


def map_header(header):
    """Map raw source header cells to target schema names via HEADER_ALIASES.
    Unrecognised columns are dropped (with a warning); missing target
    columns are filled with None later."""
    mapped = []
    for h in header:
        key = h.lower().strip()
        mapped.append(HEADER_ALIASES.get(key))  # None if unrecognised
    return mapped


def clean_wide(header, data_rows, dayfirst=True, override_site_id=None, override_source_name=None):
    mapped_header = map_header(header)

    unknown = [h for h, m in zip(header, mapped_header) if m is None]
    if unknown:
        print(f"WARNING: ignoring unrecognised source columns: {unknown}", file=sys.stderr)

    records = []
    dropped_blank = 0
    for row in data_rows:
        row_map = dict(zip(mapped_header, row))
        row_map.pop(None, None)  # drop unrecognised columns

        site_id = row_map.get("site_id") or override_site_id
        source_name = row_map.get("source_name") or override_source_name
        dt_raw = row_map.get("measurement_datetime")

        # Skip fully blank rows (no identifying info at all)
        if (site_id is None or str(site_id).strip() == "") and \
           (source_name is None or str(source_name).strip() == "") and \
           (dt_raw is None or str(dt_raw).strip() == ""):
            dropped_blank += 1
            continue

        real_dt = fix_datetime(dt_raw, dayfirst=dayfirst)

        record = {col: row_map.get(col) for col in TARGET_COLUMNS}
        record["site_id"] = int(float(site_id)) if site_id not in (None, "") else None
        record["source_name"] = source_name
        record["measurement_datetime"] = real_dt.strftime("%Y/%m/%d") if real_dt else None
        records.append(record)

    print(f"Dropped fully-blank rows: {dropped_blank}")
    return pd.DataFrame(records, columns=TARGET_COLUMNS)


def clean_long(header, data_rows, dayfirst=False):
    """Pivot a long/tidy (station, parameter, datetime, value) export into
    the wide target schema. dayfirst defaults to False here because WMIS
    exports were confirmed to use US-style M/D/Y text dates (verified by
    checking that dates increment day-by-day within a month)."""
    df = pd.DataFrame(data_rows, columns=header)
    df.columns = [c.strip() for c in df.columns]

    known_params = [p for p in df["parameter"].unique() if p.strip().lower() in PARAMETER_ALIASES]
    dropped_params = sorted(set(df["parameter"].unique()) - set(known_params))
    if dropped_params:
        print(f"WARNING: dropping parameters with no target-schema column: {dropped_params}", file=sys.stderr)

    df = df[df["parameter"].str.strip().str.lower().isin(PARAMETER_ALIASES)].copy()
    df["target_col"] = df["parameter"].str.strip().str.lower().map(PARAMETER_ALIASES)
    df["real_dt"] = df["datetime"].apply(lambda v: fix_datetime(v, dayfirst=dayfirst))
    df["measurement_datetime"] = df["real_dt"].apply(lambda d: d.strftime("%Y/%m/%d") if d else None)
    df["value"] = pd.to_numeric(df["value"], errors="coerce")

    pivot = df.pivot_table(
        index=["station", "station_name", "measurement_datetime"],
        columns="target_col",
        values="value",
        aggfunc="first",
    ).reset_index()

    pivot = pivot.rename(columns={"station": "site_id", "station_name": "source_name"})
    pivot["site_id"] = pivot["site_id"].astype(int)

    for col in TARGET_COLUMNS:
        if col not in pivot.columns:
            pivot[col] = None

    return pivot[TARGET_COLUMNS]


def clean(path: Path, sheet_name: str | None, dayfirst: bool, site_id: str | None = None, source_name: str | None = None):
    header, data_rows = load_raw_rows(path, sheet_name)
    norm_header = {h.strip().lower() for h in header}
    is_long = {"parameter", "value"}.issubset(norm_header)
    has_id_cols = bool({"site_id", "station"} & norm_header)

    if is_long:
        print("Detected long/tidy format (parameter + value columns) -> pivoting.")
        df = clean_long(header, data_rows, dayfirst=dayfirst if dayfirst is not None else False)
    else:
        if not has_id_cols:
            print("Detected 'bare wide' format (date + parameter columns, no site_id/source_name in file).")
            if not site_id or not source_name:
                sys.exit(
                    "ERROR: this file has no site_id/source_name column, so they must be supplied "
                    "with --site-id and --source-name (this metadata doesn't exist in the raw file)."
                )
        df = clean_wide(
            header, data_rows, dayfirst=dayfirst if dayfirst is not None else True,
            override_site_id=site_id, override_source_name=source_name,
        )

    # --- QA checks ---
    print(f"Total cleaned rows: {len(df)}")
    if "site_id" in df:
        print("Rows per site_id:")
        print(df["site_id"].value_counts().to_string())

    dupmask = df.duplicated(subset=["site_id", "measurement_datetime"], keep=False)
    print(f"Duplicate site_id + measurement_datetime combos after date fix: {dupmask.sum()}")
    if dupmask.sum():
        print(df[dupmask].sort_values(["site_id", "measurement_datetime"]).to_string())

    # Sort chronologically within each site
    df["_sort_dt"] = pd.to_datetime(df["measurement_datetime"], format="%Y/%m/%d", errors="coerce")
    df = df.sort_values(["site_id", "_sort_dt"]).drop(columns="_sort_dt").reset_index(drop=True)

    return df


def sanitize_name(site_id, source_name) -> str:
    """Build a filename stem like '401201_MURRAY_RIVER_JINGELLIC', matching
    the reference schema file's naming convention."""
    name = re.sub(r"[^A-Za-z0-9]+", "_", str(source_name)).strip("_").upper()
    return f"{site_id}_{name}"


def split_by_site(df: pd.DataFrame, outdir: Path) -> list[Path]:
    outdir.mkdir(parents=True, exist_ok=True)
    written = []
    for (site_id, source_name), group in df.groupby(["site_id", "source_name"], sort=True):
        stem = sanitize_name(site_id, source_name)
        out_path = outdir / f"{stem}_water_quality_cleaned.csv"
        group.to_csv(out_path, index=False)
        written.append(out_path)
    return written


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", type=Path, help="Path to the raw .xlsx or .csv file")
    parser.add_argument("--sheet", default=None, help="Sheet name to read (xlsx only, default: first sheet)")
    parser.add_argument("--out", type=Path, default=None, help="Output CSV path (combined mode)")
    parser.add_argument(
        "--split-by-site", action="store_true",
        help="Write one CSV per site_id/source_name (river/station) instead of a single combined file",
    )
    parser.add_argument(
        "--outdir", type=Path, default=None,
        help="Directory for split output files (used with --split-by-site; default: '<input_stem>_by_site' next to the input)",
    )
    parser.add_argument(
        "--dayfirst", dest="dayfirst", action="store_true", default=None,
        help="Force D/M/Y parsing for ambiguous slash-separated text dates",
    )
    parser.add_argument(
        "--monthfirst", dest="dayfirst", action="store_false",
        help="Force M/D/Y parsing for ambiguous slash-separated text dates",
    )
    parser.add_argument(
        "--site-id", default=None,
        help="site_id to apply to every row (required for 'bare wide' files with no site_id column)",
    )
    parser.add_argument(
        "--source-name", default=None,
        help="source_name to apply to every row (required for 'bare wide' files with no source_name column)",
    )
    args = parser.parse_args()

    df = clean(args.input, args.sheet, args.dayfirst, site_id=args.site_id, source_name=args.source_name)

    if args.split_by_site:
        outdir = args.outdir or args.input.with_name(args.input.stem + "_by_site")
        written = split_by_site(df, outdir)
        print(f"\nWrote {len(written)} per-site files to {outdir}/:")
        for p in written:
            print(f"  {p.name}")
    else:
        out_path = args.out or args.input.with_name(args.input.stem + "_cleaned.csv")
        df.to_csv(out_path, index=False)
        print(f"\nWrote {len(df)} rows to {out_path}")


if __name__ == "__main__":
    main()
